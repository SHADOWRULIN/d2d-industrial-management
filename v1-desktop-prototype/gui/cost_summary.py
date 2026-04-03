import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
import sqlite3
import os
import openpyxl
from openpyxl.utils import get_column_letter
from database.db_connection import connect_db  

class CostSummaryInterface:
    def __init__(self, root, proposal_id):
        self.proposal_id = proposal_id
        self.root = root
        self.root.title(f"Cost Summary - Proposal ID {proposal_id}")
        self.root.geometry("900x600")

        tk.Label(self.root, text="Vendor Bill Cost Summary", font=("Arial", 18)).pack(pady=10)

        # Treeview Table
        self.tree = ttk.Treeview(
            self.root,
            columns=("SNo", "BillNo", "Vendor", "Phone", "VendorType", "TotalAmount", "PaidAmount", "Balance", "PaymentDate"),
            show="headings"
        )
        for col, label in zip(
            self.tree["columns"],
            ["S.No.", "Bill No", "Vendor Name", "Phone Number", "Vendor Type", "Total Amount", "Paid Amount", "Balance", "Payment Date"]
        ):
            self.tree.heading(col, text=label)
            self.tree.column(col, width=120)
        self.tree.column("SNo", width=50)
        self.tree.pack(fill="both", expand=True, pady=10)

        # Buttons
        tk.Button(self.root, text="Calculate Total Cost", command=self.calculate_total).pack(pady=5)
        tk.Button(self.root, text="Export to Excel", command=self.export_to_excel).pack(pady=5)

        self.load_vendor_data()

    def load_vendor_data(self):
        conn = connect_db()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT v.vendor_name, v.vendor_phone, v.vendor_type, SUM(v.quantity * v.price) as total_amount, 
                   p.payment_date, COALESCE(SUM(p.amount), 0) as paid_amount
            FROM vendors v
            LEFT JOIN payments p ON v.vendor_name = p.vendor_name AND v.proposal_id = p.proposal_id
            WHERE v.proposal_id = ?
            GROUP BY v.vendor_name, v.vendor_phone, v.vendor_type, p.payment_date
        """, (self.proposal_id,))
        raw_data = cursor.fetchall()
        conn.close()

        self.vendor_data = []
        self.tree.delete(*self.tree.get_children())

        for idx, (name, phone, vtype, total_amount, payment_date, paid_amount) in enumerate(raw_data, start=1):
            bill_no = f"BILL{idx:03d}"
            balance = total_amount - paid_amount
            self.vendor_data.append((bill_no, name, phone, vtype, total_amount, paid_amount, balance, payment_date))
            self.tree.insert("", "end", values=(idx, bill_no, name, phone, vtype, f"{total_amount:.2f}", f"{paid_amount:.2f}", f"{balance:.2f}", payment_date))

    def calculate_total(self):
        total_cost = total_paid = total_balance = 0.0
        for row in self.tree.get_children():
            values = self.tree.item(row)["values"]
            total_cost += float(values[5])  
            total_paid += float(values[6])
            total_balance += float(values[7])

        messagebox.showinfo("Total Cost Summary", f"Total Amount: Rs. {total_cost:.2f}\nTotal Paid: Rs. {total_paid:.2f}\nRemaining Balance: Rs. {total_balance:.2f}")

    def export_to_excel(self):
        if not self.vendor_data:
            messagebox.showwarning("No Data", "No data to export.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save Cost Summary as Excel"
        )
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Vendor Cost Summary"

            headers = ["S.No.", "Bill No", "Vendor Name", "Phone Number", "Vendor Type", "Total Amount", "Paid Amount", "Balance", "Payment Date"]
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)

            for idx, (bill_no, name, phone, vtype, total_amount, paid_amount, balance, payment_date) in enumerate(self.vendor_data, start=1):
                ws.cell(row=idx+1, column=1, value=idx)
                ws.cell(row=idx+1, column=2, value=bill_no)
                ws.cell(row=idx+1, column=3, value=name)
                ws.cell(row=idx+1, column=4, value=phone)
                ws.cell(row=idx+1, column=5, value=vtype)
                ws.cell(row=idx+1, column=6, value=round(total_amount, 2))
                ws.cell(row=idx+1, column=7, value=round(paid_amount, 2))
                ws.cell(row=idx+1, column=8, value=round(balance, 2))
                ws.cell(row=idx+1, column=9, value=payment_date)

            for col_num in range(1, 10):
                col_letter = get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = 20

            wb.save(file_path)
            messagebox.showinfo("Exported", f"Excel file saved successfully at:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export to Excel:\n{e}")

if __name__ == "__main__":
    root = tk.Tk()
    CostSummaryInterface(root, "1")
    root.mainloop()
