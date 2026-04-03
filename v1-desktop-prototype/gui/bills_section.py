import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
import sqlite3
import openpyxl
from openpyxl.utils import get_column_letter
from database.db_connection import connect_db  

class BillsSection:
    def __init__(self, root, proposal_id):
        self.proposal_id = proposal_id
        self.root = root
        self.root.title(f"Bills Management - Proposal ID {proposal_id}")
        self.root.geometry("900x600")

        self.bill_counter = 0  # For generating unique bill numbers

        tk.Label(self.root, text=f"Bills for Proposal ID {proposal_id}", font=("Arial", 18)).pack(pady=10)

        # Vendor Type Dropdown
        tk.Label(self.root, text="Select Vendor Type:").pack()
        self.vendor_type_var = tk.StringVar(self.root)
        self.vendor_type_var.set("Raw Material")
        vendor_options = ["Raw Material", "Parts/Items"]
        tk.OptionMenu(self.root, self.vendor_type_var, *vendor_options).pack()

        # Vendor Name Dropdown
        tk.Label(self.root, text="Select Vendor Name (optional):").pack()
        self.vendor_name_var = tk.StringVar()
        self.vendor_name_combo = ttk.Combobox(self.root, textvariable=self.vendor_name_var, postcommand=self.load_vendor_names)
        self.vendor_name_combo.pack()

        # Vendor Phone Entry (optional)
        tk.Label(self.root, text="Enter Vendor Phone Number (optional):").pack()
        self.vendor_phone_entry = tk.Entry(self.root)
        self.vendor_phone_entry.pack()

        # Date Selection Dropdown
        tk.Label(self.root, text="Select Payment Date:").pack()
        self.payment_date_entry = DateEntry(self.root, width=12, background='darkblue', foreground='white', borderwidth=2)
        self.payment_date_entry.pack()

        # Buttons
        tk.Button(self.root, text="Retrieve Vendor Bills", command=self.retrieve_invoices).pack(pady=10)
        tk.Button(self.root, text="Export to Excel", command=self.export_to_excel).pack(pady=5)

        # Bill Number Display
        self.bill_label = tk.Label(self.root, text="Latest Bill No: N/A", font=("Arial", 12, "bold"), fg="blue")
        self.bill_label.pack()

        # Table with Payment Date Column
        self.tree = ttk.Treeview(self.root, columns=("SNo", "Material", "Qty", "PricePerUnit", "Total", "PaymentDate"), show="headings")
        for col, label in zip(self.tree["columns"], ["S.No.", "Material", "Qty", "Price/Unit", "Total Price", "Payment Date"]):
            self.tree.heading(col, text=label)
            self.tree.column(col, width=120)
        self.tree.column("SNo", width=50)
        self.tree.pack(fill="both", expand=True, pady=10)

    def generate_bill_number(self):
        self.bill_counter += 1
        return f"BILL{self.bill_counter:03d}"

    def load_vendor_names(self):
        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT vendor_name FROM vendors WHERE proposal_id = ?", (self.proposal_id,))
            vendors = [row[0] for row in cursor.fetchall()]
            conn.close()
            self.vendor_name_combo['values'] = vendors
        except Exception as e:
            messagebox.showerror("Error", f"Could not load vendor names:\n{e}")

    def retrieve_invoices(self):
        vendor_type = self.vendor_type_var.get()
        vendor_name = self.vendor_name_var.get().strip()
        vendor_phone = self.vendor_phone_entry.get().strip()
        payment_date = self.payment_date_entry.get_date()

        query = "SELECT material, quantity, price, purchase_date FROM vendors WHERE proposal_id = ?"
        params = [self.proposal_id]

        if vendor_type:
            query += " AND vendor_type = ?"
            params.append(vendor_type)
        if vendor_name:
            query += " AND vendor_name = ?"
            params.append(vendor_name)
        if vendor_phone:
            query += " AND vendor_phone = ?"
            params.append(vendor_phone)
        if payment_date:
            query += " AND purchase_date = ?"
            params.append(payment_date)

        # Clear the table before loading
        for i in self.tree.get_children():
            self.tree.delete(i)

        self.bill_counter = 0

        try:
            conn = connect_db()
            cursor = conn.cursor()
            cursor.execute(query, tuple(params))
            data = cursor.fetchall()
            conn.close()

            if not data:
                self.bill_label.config(text="Latest Bill No: N/A")
                messagebox.showinfo("No Data", "No matching vendor bills found for the selected criteria.")
                return

            for idx, row in enumerate(data, start=1):
                material, qty, unit_price, purchase_date = row
                total_price = float(qty) * float(unit_price)
                bill_no = self.generate_bill_number()
                self.tree.insert('', 'end', values=(idx, material, qty, unit_price, total_price, purchase_date))
                self.bill_label.config(text=f"Latest Bill No: {bill_no}")

        except Exception as e:
            messagebox.showerror("Database Error", f"Error while retrieving bills:\n{e}")

    def export_to_excel(self):
        if not self.tree.get_children():
            messagebox.showwarning("Warning", "No data to export!")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save Excel File"
        )
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Proposal {self.proposal_id} Bills"

            headers = ["S.No.", "Material", "Quantity", "Price/Unit", "Total Price", "Payment Date"]
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)

            for row_num, item_id in enumerate(self.tree.get_children(), start=2):
                row_data = self.tree.item(item_id)["values"]
                for col_num, value in enumerate(row_data, start=1):
                    ws.cell(row=row_num, column=col_num, value=value)

            for col_num in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = 15

            wb.save(file_path)
            messagebox.showinfo("Success", f"Excel file saved successfully:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Export Error", f"An error occurred:\n{str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    BillsSection(root, "1")
    root.mainloop()
